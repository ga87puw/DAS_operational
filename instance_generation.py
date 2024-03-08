#!/usr/bin/env python3


import networkx as nx
import random
from scipy.spatial import distance
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import math
import openpyxl
import time
import os
import pickle
import gurobi_model
import grasp_heuristic
import heuristic_path


def min_dist(n, m, d):
    """
    Check for a stop whether the minimum distance to other stops is respected.

    Parameters:
    n (int, int): stop to check
    m (list of (int, int)): other existing stops
    d (int): minimum distance

    Returns:
    boolean: True if minimum distance is respected, False otherwise
    """
    for x in m:
        if distance.cityblock(n, x) < d:
            return False
    return True


def max_dist(n, m, d):
    """
    Check for a request location whether there is at least one stop within a maximum distance.

    Parameters:
    n (int, int): request location to check
    m (list of (int, int)): existing stops
    d (int): maximum distance

    Returns:
    boolean: True if there is a stop within the maximum distance, False otherwise
    """
    for x in m:
        if distance.cityblock(n, x) <= d:
            return x
    return False


def stops_in_walking_distance(n, m, d, loc=True):
    """
    Return all stops within the maximum walking willingness of a location.

    Parameters:
    n (int, int): request location to check
    m (list of (int, int)): existing stops
    d (int): maximum walking willingness
    loc (bool): True if location requests are used, False indicates stop requests

    Returns:
    list of (int, int): List of stops within walking willingness (for stop requests: closest stop)
    """
    t = []

    # all stops
    if loc:
        for k, v in m.items():
            if distance.cityblock(n, v) <= d:
                t.append(k)

    # best stop
    else:
        best = 1000000
        for k, v in m.items():
            if distance.cityblock(n, v) <= d and distance.cityblock(n, v) < best:
                best = distance.cityblock(n, v)
                t = [k]
    return t


def request_feasible(n1, n2, m, d):
    """
    Check if the request is viable (every check except the check of orientation).

    Parameters:
    n1 (int, int): request origin to check
    n1 (int, int): request destination to check
    m (list of (int, int)): existing stops
    d (int): maximum distance

    Returns:
    boolean: True if the request is viable, False otherwise
    """
    m_list = m.values()

    # at least one stop in reach
    if not max_dist(n1, m_list, d) or not max_dist(n2, m_list, d):
        return False

    # not directly at a stop
    if max_dist(n1, m_list, 2) or max_dist(n2, m_list, 2):
        return False

    # minimum horizontal distance
    if not abs(n1[0] - n2[0]) > 3 * d:
        return False

    # no optional stops in the same segment for boarding and alighting
    t1 = stops_in_walking_distance(n1, m, d)
    t2 = stops_in_walking_distance(n2, m, d)
    t1b = [grasp_heuristic.segment(s) for s in t1 if s[0] == 'o']
    t2b = [grasp_heuristic.segment(s) for s in t2 if s[0] == 'o']
    return not any(seg in t1b for seg in t2b)


def data_gen(n_compulsory, n_optional, n_requests, time_window_width, seed, vis_graph=False, loc=True):
    """
    Create an operational DAS problem instance.

    Parameters:
    n_compulsory (int): number of compulsory stops
    n_optional (int): number of optional stops
    n_requests (int): number of requests
    time_window_width (int): time window width in seconds
    seed (int): random seed
    vis_graph (bool): indicates if graph should be visualized, only useful for small instances
    loc (bool): True for location requests, False for stop requests

    Returns:
    str: filename of pickle file for this instance (stored in folder: 'data')
    """
    #######################################################################################################
    # n_compulsory = 8  # number of compulsory stops
    # n_optional = 4  # number of optional stops per segment
    # n_requests = 24  # number of requests
    # time_window_width = 180  # length of the time window in seconds
    time_window_factor = 1.5  # segment time as a multiple of basic path duration
    height = 15  # distance nodes can have at most to the center line (overall height: x2) in distance units
    length = 50  # approximately the length between two compulsory stops in distance units

    # the bus has an average speed of 20 km/h, users walking speed is on average 4 km/h. -> bus 5 times as fast as users
    bus_multi = 4  # how many seconds it takes the bus to travel 1 distance units
    user_multi = 20  # how many seconds it takes the bus to travel 1 distance units
    min_distance = 5  # minimum distance between two stops -> 20 seconds
    max_distance = 12  # maximum distance customers are willing to walk to and from bus stops -> 240 seconds

    stop_time = 10  # how many seconds it takes the bus to stop at a stop

    random.seed(seed)
    #######################################################################################################
    data_gen_runtime = time.time()

    G = nx.DiGraph()  # whole graph
    G_c = nx.Graph()  # compulsory stops
    G_h = [nx.DiGraph() for i in range(n_compulsory - 1)]  # graph segments
    G_b = nx.DiGraph()  # bus related graph
    G_s = nx.Graph()  # origins
    G_e = nx.Graph()  # destinations

    # === Step 1: Creating stops ===
    # terminal
    G_c.add_node('c00', pos=(0, 0))
    tmp_node = (0, 0)
    pos_list = [(0, 0)]
    for i in range(n_compulsory - 1):
        G_h[i].add_node(f'c{i:02d}', pos=tmp_node)

        # compulsory stop
        b = (i + 1) * length, 0

        pos_list.append(b)
        G_c.add_node(f'c{i + 1:02d}', pos=b)
        G_h[i].add_node(f'c{i + 1:02d}', pos=b)

        # optional stops
        seg_ellipse = patches.Ellipse((b[0] - length / 2, 0), length, 2 * height + 5)

        for j in range(n_optional):
            c = (random.randint(tmp_node[0] + 1, b[0] - 1), random.randint(-height, height))

            # ensure min distance to any other stop and that the stop is within its segment
            while not min_dist(c, pos_list, min_distance) or not seg_ellipse.contains_point(c):
                c = (random.randint(tmp_node[0] + 1, b[0] - 1), random.randint(-height, height))

            pos_list.append(c)
            G_h[i].add_node(f'o{i:02d},{j:02d}', pos=c)

        # add edges
        for n1 in G_h[i]:
            for n2 in G_h[i]:
                if n1 != f'c{i + 1:02d}' and n2 != f'c{i:02d}' and n1 != n2:
                    G_h[i].add_edge(n1, n2, weight=round(
                        distance.cityblock(G_h[i].nodes[n1]['pos'], G_h[i].nodes[n2]['pos'])) * bus_multi + stop_time)

        # update graph G
        G = nx.compose(G, G_h[i])

        # remove not needed positions
        pos_list = pos_list[-n_optional - 1:]
        tmp_node = b

    G_b = G.copy()
    pos_stops_dict = nx.get_node_attributes(G, 'pos')

    # === Step 2: Creating requests ===
    for i in range(n_requests):

        # start and end nodes of requests
        s = (random.randint(0, (n_compulsory - 1) * length), random.randint(-height, height))
        e = (random.randint(0, (n_compulsory - 1) * length), random.randint(-height, height))

        # repeat until there is a horizontal min distance between s and e and both are within reach of stops
        while not request_feasible(s, e, pos_stops_dict, max_distance):
            s = (random.randint(0, (n_compulsory - 1) * length), random.randint(-height, height))
            e = (random.randint(0, (n_compulsory - 1) * length), random.randint(-height, height))

        # s has to be to the left of e
        if s[0] > e[0]:
            s, e = e, s

        # add nodes
        G_s.add_node(f's{i:02d}', pos=s)
        G_e.add_node(f'e{i:02d}', pos=e)
        G.add_node(f's{i:02d}', pos=s)
        G.add_node(f'e{i:02d}', pos=e)

        # add edges
        for n2 in stops_in_walking_distance(s, pos_stops_dict, max_distance, loc=loc):
            G.add_edge(f's{i:02d}', n2, weight=round(distance.cityblock(s, G.nodes[n2]['pos']) * user_multi))
        for n1 in stops_in_walking_distance(e, pos_stops_dict, max_distance, loc=loc):
            G.add_edge(n1, f'e{i:02d}', weight=round(distance.cityblock(G.nodes[n1]['pos'], e) * user_multi))

        # edge for not using the bus (penalty): fixed part that depends also on the number of requests
        #                                   + variable part that depends on the distance between origin and destination
        G.add_edge(f's{i:02d}', f'e{i:02d}',
                   weight=1000 + 100 * n_requests + round(distance.cityblock(s, e) * bus_multi * time_window_factor))

    # === Step 3: Visualization ===
    if vis_graph:
        pos = nx.get_node_attributes(G, 'pos')
        plt.figure(figsize=(((n_compulsory - 1) * length) / 8, height / 8))

        # add segments
        for n_seg in range(n_compulsory - 1):
            ellipse = patches.Ellipse(((n_seg + 0.5) * length, 0), length, height * 2 + 5,
                                      angle=0, edgecolor=(0, 0, 0, 0.4), facecolor='none')
            plt.gca().add_patch(ellipse)

        # visualize stops and requests
        l1 = nx.draw_networkx_nodes(G_c, pos, node_size=144, node_color='k', node_shape='s', label='compulsory stop')
        l2 = nx.draw_networkx_nodes(set(G.nodes()) - set(G_c.nodes()) - set(G_s.nodes()) - set(G_e.nodes()), pos,
                                    node_size=44, node_color='k', label='optional stops')
        l3 = nx.draw_networkx_nodes(G_s, pos, node_size=44, node_color='b', node_shape='^', label='user origins')
        l4 = nx.draw_networkx_nodes(G_e, pos, node_size=44, node_color='g', node_shape='^', label='user destinations')

        plt.legend(handles=[l1, l2, l3, l4], loc='upper right')
        plt.savefig('graph.png')
        plt.show()

    # === Step 4: Creating remaining data ===
    # requests
    requests = [f'r{i:02d}' for i in range(n_requests)]
    # demand
    demand = {x: math.ceil(random.expovariate(2)) for x in requests}
    # time windows
    a_time_window = [0]
    b_time_window = [0]
    for i in range(n_compulsory - 1):
        x = G.get_edge_data(f'c{i:02d}', f'c{i + 1:02d}')['weight']
        a_time_window.append(round(math.ceil(a_time_window[-1] + x * time_window_factor)))
        b_time_window.append(round(a_time_window[-1] + time_window_width))

    # big M
    big_M = b_time_window[-1]

    # create instance
    instance = gurobi_model.Instance(G, G_b, G_c, G_s, G_e, G_h, demand, requests, a_time_window, b_time_window, big_M)

    data_gen_runtime = time.time() - data_gen_runtime
    print(f'It took {data_gen_runtime} seconds to create the data')

    # filename
    filename = f'instance_{n_compulsory}_{n_optional}_{n_requests}_{time_window_factor}_{time_window_width}_{seed}'
    # if stop requests instead of location requests
    if not loc:
        filename = filename + '_loc'

    # save instance to 'data' folder
    file_path = os.path.join('data', filename)
    with open(file_path, 'wb') as file:
        pickle.dump(instance, file)

    return filename


def opt_or_heu(filename, seed_instance, heuristic, heuristic_factor=None, vis_res=False, vis_perf=False):
    """
    Solve an operational DAS problem instance. Also store results to an excel file

    Parameters:
    filename (str): filename of the instance in folder 'data'
    seed_instance (int): random seed that was used for the instance creation
    heuristic (int): if 0, MILP benchmark, 3, solution approach with arc-based, 4, solution approach with path-based
    heuristic_factor (int, int, int): (completion bonus, candidate list length)
    vis_res (bool): indicates if a visualization of the should be produced, only useful for small instances

    Returns:
    str: filename of pickle file for this instance (stored in folder: 'data')
    """
    # load instance
    file_path = os.path.join('data', filename)
    with open(file_path, 'rb') as file:
        instance = pickle.load(file)
    n_compulsory = len(instance.C)
    n_optional = instance.G_h[0].number_of_nodes() - 2
    n_requests = len(instance.R)
    time_window_width = instance.b[1] - instance.a[1]

    # === Step 5: Optimization or heuristic (+ optimization) ===
    heuristic_runtime = time.time()
    if not heuristic:
        heuristic_name = 'MILP benchmark'
        result = gurobi_model.optimize_arc(instance)
        heuristic_runtime = 0
    elif heuristic == 3:
        heuristic_name = 'arc-based grasp'
        # find route via heuristic
        result, _ = heuristic.grasp(instance, heuristic_factor)
        heuristic_runtime = time.time() - heuristic_runtime
        print(f'It took {heuristic_runtime} seconds to perform the heuristic {heuristic_name}')

    elif heuristic == 4:
        heuristic_name = 'path-based grasp'
        # find route via heuristic
        result, _ = heuristic_path.grasp_for_path_generation(instance, heuristic_factor)
        heuristic_runtime = time.time() - heuristic_runtime
        print(f'It took {heuristic_runtime} seconds to perform the heuristic {heuristic_name}')

    # === Step 6: Visualizing and/or storing results===
    if result:
        bus_route, walks, denied, penalty, cur_p, tour_time, waiting, parts, performance, final_result = result

        # visualize results
        if vis_res:
            pos = nx.get_node_attributes(instance.G, 'pos')

            # for visualization purposes
            height = 15  # this is the distance nodes can have at most to the center line (overall height: x2)
            length = 50  # approximately the length between two compulsory stops
            plt.figure(figsize=(((n_compulsory - 1) * length) / 8, height / 8))

            plt.figure()

            # add segments
            '''for n_seg in range(n_compulsory - 1):
                ellipse = patches.Ellipse(((n_seg + 0.5) * length, 0), length, height * 2 + 5,
                                          angle=0, edgecolor=(0, 0, 0, 0.4), facecolor='none')
                plt.gca().add_patch(ellipse)'''

            nx.draw_networkx_nodes(instance.C, pos, node_size=44, node_color='k', node_shape='s')
            nx.draw_networkx_nodes(instance.N_S, pos, node_size=4, node_color='b', node_shape='^')
            nx.draw_networkx_nodes(instance.N_E, pos, node_size=4, node_color='g', node_shape='^')
            nx.draw_networkx_nodes(set(instance.G.nodes()) - set(instance.C) - set(instance.N_S) - set(instance.N_E),
                                   pos, node_size=4, node_color='k')
            nx.draw_networkx_edges(instance.G, pos, edgelist=bus_route, edge_color='k', width=1, alpha=0.8)
            nx.draw_networkx_edges(instance.G, pos, edgelist=walks[0], edge_color='b', width=0.5, alpha=0.5)
            nx.draw_networkx_edges(instance.G, pos, edgelist=walks[1], edge_color='g', width=0.5, alpha=0.5)
            nx.draw_networkx_edges(instance.G, pos, edgelist=walks[2], edge_color='r', width=0.5, alpha=0.5)

            # plt.savefig('graph_solved.png')
            plt.show()

        if vis_perf:
            # visualize performance
            # extract the data for plotting
            times = [entry[0] for entry in performance]
            objective_values = [entry[1] for entry in performance]
            best_bounds = [entry[2] for entry in performance]

            # create a plot
            plt.clf()
            plt.figure(figsize=(10, 6))
            plt.plot(times, objective_values, label='Objective Value')
            plt.plot(times, best_bounds, label='Best Bound')
            plt.xlabel('Time (seconds)')
            plt.ylabel('Value')
            plt.title(f'comp: {n_compulsory}, opt: {n_optional}, req: {n_requests}, seed: {seed_instance}')
            plt.legend()

            plt.show()

        # parts: ('walking, driving(/+w), penalty, (/waiting)')
        if len(parts) == 3:
            parts = (*parts, 0)

        n_users = sum(instance.d.values())
        n_users_accepted = n_users - denied

        # results to Excel
        result_dict = {"heuristic": heuristic_name,
                       "heuristic_factor": str(heuristic_factor),
                       "n_compulsory": n_compulsory,
                       "n_optional": n_optional,
                       "n_requests": n_requests,
                       "time_window": time_window_width,
                       "seed": seed_instance,
                       "name": filename,

                       "heuristic runtime": round(heuristic_runtime, 2),
                       "solver runtime": round(final_result[0], 2),
                       "objective value": final_result[1],  # upper bound
                       "best bound": final_result[2],  # lower bound
                       "gap": round(final_result[3], 4),

                       "penalty": penalty,
                       "not penalty": final_result[1] - penalty,
                       'walking': parts[0],
                       'sitting (in bus)': parts[1] + parts[3],
                       "denied": denied,
                       "percentage denied": round(denied / n_users, 4),
                       }

        # open the Excel file (create if it doesn't exist)
        excel_name = f"output{n_compulsory}_{n_optional}_{n_requests}_{time_window_width}.xlsx"

        # Check if the file exists before loading it
        if os.path.isfile(excel_name):
            workbook = openpyxl.load_workbook(excel_name)
        else:
            workbook = openpyxl.Workbook()
        sheet = workbook.active
        # write the data to the Excel sheet
        if sheet.max_row == 1:
            # add header row if it's a new file
            sheet.append(list(result_dict.keys()))

        sheet.append(list(result_dict.values()))
        # save the Excel file
        workbook.save(excel_name)


def compare_approaches(n_compulsory, n_optional, n_requests, time_window_factor, time_window_width, seeds):
    """
    Create multiple instances and run the different solution approach variants for them.

    Parameters:
    n_compulsory (int): number of compulsory stops
    n_optional (int): number of optional stops
    n_requests (int): number of requests
    time_window_factor (int): time window factor
    time_window_width (int): time window width in seconds
    seeds (list(int)): random seeds

    Saves results as excel file
    """
    # create instances
    for seed in seeds:
        data_gen(n_compulsory, n_optional, n_requests, time_window_width, seed)

    for seed in seeds:
        data = f'instance_{n_compulsory}_{n_optional}_{n_requests}_{time_window_factor}_{time_window_width}_{seed}'
        # MILP benchmark
        opt_or_heu(data, seed, heuristic=0)

    for seed in seeds:
        data = f'instance_{n_compulsory}_{n_optional}_{n_requests}_{time_window_factor}_{time_window_width}_{seed}'
        # A1
        opt_or_heu(data, seed, heuristic=3, heuristic_factor=[1, 1, 1])

    for seed in seeds:
        data = f'instance_{n_compulsory}_{n_optional}_{n_requests}_{time_window_factor}_{time_window_width}_{seed}'
        # A100
        opt_or_heu(data, seed, heuristic=3, heuristic_factor=[1, 5, 100])

    for seed in seeds:
        data = f'instance_{n_compulsory}_{n_optional}_{n_requests}_{time_window_factor}_{time_window_width}_{seed}'
        # P 100
        opt_or_heu(data, seed, heuristic=4, heuristic_factor=[1, 5, 100])


if __name__ == "__main__":
    # (compulsory stops, optional stops, requests, time window factor, time window width, list of seeds)
    compare_approaches(5, 10, 10, 1.5, 0, list(range(10)))
